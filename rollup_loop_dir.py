#!/usr/local/bin/python3

import string
import matplotlib.pyplot as plt

# https://automatetheboringstuff.com/chapter12/

# import load_workbook
from openpyxl import load_workbook


import os

teams = ['DAPS', 'DBA' , 'INF', 'RCI', 'COLLAB', 'MMS', 'TNS-V', 'TNS-NE', 'TNS-FS','PS-OTHER', 'TEST' ]

team  = input("What team do ya want to rollup? ")
if team.upper() not in teams:
   print ("Not a Valid Team!")
   print (teams)
   quit()

print()
print("If there are errors, most likey there are xlxs files besides the RM Workbooks")
print()

#scratch_dir = "/Users/jeffreywiedemann/Desktop/Resource_Planning/Team_Reports/" + team.upper() + "/"
scratch_dir = "/DATA/" + team.upper() + "/"
#csvfile = open("/Users/jeffreywiedemann/Desktop/Resource_Planning/Team_Reports/" + team.upper() + "/" + team.lower() + "_rollup.csv","w+")
csvfile = open("/DATA/" + team.upper() + "/" + team.lower() + "_rollup.csv","w+")


print("Name,Department,Total Time Away,Total General Admin,Total Managerial Admin Time,TOTAL Admin Time,Total Support,Total Consulting,Total Other,Annual CI Project Hours,Annual AS Project Hours,Annual IT SS Project Hours,Annual ISO Project Hours,Annual Schools or Depts Project Hours,TOTAL Project Hours")

csvfile.write("Name,Department,Total Time Away,Total General Admin,Total Managerial Admin Time,TOTAL Admin Time,Total Support,Total Consulting,Total Other,Annual CI Project Hours,Annual AS Project Hours,Annual IT SS Project Hours,Annual ISO Project Hours,Annual Schools or Depts Project Hours,TOTAL Project Hours\n")

People = 0
Sum_Total_Time_Away = 0
Sum_Total_General_Admin = 0
Sum_Total_Managerial_Admin_Time = 0
Sum_TOTAL_Admin_Time = 0
Sum_Total_Support = 0
Sum_Total_Consulting = 0
Sum_Total_Other = 0
Sum_Annual_CI_Project_Hours = 0
Sum_Annual_AS_Project_Hours = 0
Sum_Annual_IT_SS_Project_Hours = 0
Sum_Annual_ISO_Project_Hours = 0
Sum_Annual_Schools_or_Depts_Project_Hours = 0
Sum_TOTAL_Project_Hours = 0
Sum_TOTAL_Hours = 0


for filename in os.listdir(scratch_dir):
   if filename.endswith(".xlsx"):

     #Uncomment below to see filename during testing. Testing?
     #print (scratch_dir + filename)

     # set file path
     #filepath="/Users/jeffreywiedemann/Desktop/Resource_Planning/jeff_for_python.xlsx"
     filepath = scratch_dir + filename
     
     wb=load_workbook(filepath, data_only=True)
     
     sheet = wb['Summary']
     
     Name = sheet['A2']
     Dept = sheet['A1']
     
     Total_Time_Away = sheet['C16']
     Total_General_Admin = sheet['C22']
     Total_Managerial_Admin_Time = sheet['C25']
     TOTAL_Admin_Time = sheet['C16'].value + sheet['C22'].value + sheet['C25'].value
     Total_Support = sheet['C33']
     Total_Consulting = sheet['C37']
     Total_Other = sheet['C41'] 

     Annual_CI_Project_Hours = sheet['C50']
     Annual_AS_Project_Hours = sheet['C51']
     Annual_IT_SS_Project_Hours = sheet['C52']
     Annual_ISO_Project_Hours = sheet['C53']
     Annual_Schools_or_Depts_Project_Hours = sheet['C54']
     TOTAL_Project_Hours = sheet['C44']

     TOTAL_Hours = sheet['C16'].value +\
     sheet['C22'].value +\
     sheet['C25'].value +\
     sheet['C33'].value +\
     sheet['C37'].value +\
     sheet['C41'].value +\
     sheet['C50'].value +\
     sheet['C51'].value +\
     sheet['C52'].value +\
     sheet['C53'].value +\
     sheet['C54'].value



     Sum_Total_Time_Away = Sum_Total_Time_Away + Total_Time_Away.value
     Sum_Total_General_Admin = Sum_Total_General_Admin + Total_General_Admin.value
     Sum_Total_Managerial_Admin_Time = Sum_Total_Managerial_Admin_Time + Total_Managerial_Admin_Time.value
     Sum_TOTAL_Admin_Time = Sum_TOTAL_Admin_Time + TOTAL_Admin_Time
     Sum_Total_Support = Sum_Total_Support + Total_Support.value
     Sum_Total_Consulting = Sum_Total_Consulting + Total_Consulting.value
     Sum_Total_Other = Sum_Total_Other + Total_Other.value
     Sum_Annual_CI_Project_Hours = Sum_Annual_CI_Project_Hours + Annual_CI_Project_Hours.value
     Sum_Annual_AS_Project_Hours = Sum_Annual_AS_Project_Hours + Annual_AS_Project_Hours.value
     Sum_Annual_IT_SS_Project_Hours = Sum_Annual_IT_SS_Project_Hours + Annual_IT_SS_Project_Hours.value
     Sum_Annual_ISO_Project_Hours = Sum_Annual_ISO_Project_Hours + Annual_ISO_Project_Hours.value
     Sum_Annual_Schools_or_Depts_Project_Hours = Sum_Annual_Schools_or_Depts_Project_Hours + Annual_Schools_or_Depts_Project_Hours.value
     Sum_TOTAL_Project_Hours = Sum_TOTAL_Project_Hours + TOTAL_Project_Hours.value
     Sum_TOTAL_Hours = Sum_TOTAL_Hours + TOTAL_Hours
   

     People = People + 1

     print(Name.value, end=",")
     csvfile.write(Name.value + ",")

     print(Dept.value, end=",")
     csvfile.write(Dept.value + ",")

     print(Total_Time_Away.value, end=",")
     csvfile.write(str(Total_Time_Away.value) + ",")

     print(Total_General_Admin.value, end=",")
     csvfile.write(str(Total_General_Admin.value) + ",")

     print(Total_Managerial_Admin_Time.value, end=",")
     csvfile.write(str(Total_Managerial_Admin_Time.value) + ",")

     print(TOTAL_Admin_Time, end=",")
     csvfile.write(str(TOTAL_Admin_Time) + ",")

     print(Total_Support.value, end=",")
     csvfile.write(str(Total_Support.value) + ",")

     print(Total_Consulting.value, end=",")
     csvfile.write(str(Total_Consulting.value) + ",")

     print(Total_Other.value, end=",")
     csvfile.write(str(Total_Other.value) + ",")

     print(Annual_CI_Project_Hours.value, end=",")
     csvfile.write(str(Annual_CI_Project_Hours.value) + ",")

     print(Annual_AS_Project_Hours.value, end=",")
     csvfile.write(str(Annual_AS_Project_Hours.value) + ",")

     print(Annual_IT_SS_Project_Hours.value, end=",")
     csvfile.write(str(Annual_IT_SS_Project_Hours.value) + ",")

     print(Annual_ISO_Project_Hours.value, end=",")
     csvfile.write(str(Annual_ISO_Project_Hours.value) + ",")

     print(Annual_Schools_or_Depts_Project_Hours.value, end=',')
     csvfile.write(str(Annual_Schools_or_Depts_Project_Hours.value) + ",")

     print(TOTAL_Project_Hours.value, end=',')
     csvfile.write(str(TOTAL_Project_Hours.value) + ",")

     print(TOTAL_Hours)
     csvfile.write(str(TOTAL_Hours))





# Keep these commnets beccause they illustrate working with diiferent worksheets within a workbook
#     sheet = wb['Meetings & Admin']
#
#     M_B10=sheet['B10']
#
#     print(M_B10.value)
#     csvfile.write(str(M_B10.value) + ",")

     csvfile.write("\n")


# end of for filename in os.listdir(scratch_dir):
# print(Sum_Total_Time_Away)


csvfile.write("TOTAL,,")

csvfile.write(str(Sum_Total_Time_Away) + ",")

csvfile.write(str(Sum_Total_General_Admin) + ",")

csvfile.write(str(Sum_Total_Managerial_Admin_Time) + ",")

csvfile.write(str(Sum_TOTAL_Admin_Time) + ",")

csvfile.write(str(Sum_Total_Support) + ",")

csvfile.write(str(Sum_Total_Consulting) + ",")

csvfile.write(str(Sum_Total_Other) + ",")

csvfile.write(str(Sum_Annual_CI_Project_Hours) + ",")

csvfile.write(str(Sum_Annual_AS_Project_Hours) + ",")

csvfile.write(str(Sum_Annual_IT_SS_Project_Hours) + ",")

csvfile.write(str(Sum_Annual_ISO_Project_Hours) + ",")

csvfile.write(str(Sum_Annual_Schools_or_Depts_Project_Hours) + ",")

csvfile.write(str(Sum_TOTAL_Project_Hours) + ",")

csvfile.write(str(Sum_TOTAL_Hours))

csvfile.write("\n")

The_Sums = Sum_TOTAL_Admin_Time +  Sum_Total_Support +  Sum_Total_Consulting +  Sum_Total_Other + Sum_TOTAL_Project_Hours

Percentage_Total_Time_Away = Sum_Total_Time_Away / Sum_TOTAL_Hours * 100
Percentage_Total_General_Admin = Sum_Total_General_Admin / Sum_TOTAL_Hours * 100
Percentage_Total_Managerial_Admin_Time = Sum_Total_Managerial_Admin_Time / Sum_TOTAL_Hours * 100
Percentage_TOTAL_Admin_Time = Sum_TOTAL_Admin_Time / Sum_TOTAL_Hours * 100
Percentage_Total_Support = Sum_Total_Support / Sum_TOTAL_Hours * 100
Percentage_Total_Consulting = Sum_Total_Consulting / Sum_TOTAL_Hours * 100
Percentage_Total_Other = Sum_Total_Other / Sum_TOTAL_Hours * 100
Percentage_Annual_CI_Project_Hours = Sum_Annual_CI_Project_Hours / Sum_TOTAL_Hours * 100
Percentage_Annual_AS_Project_Hours = Sum_Annual_AS_Project_Hours / Sum_TOTAL_Hours * 100
Percentage_Annual_IT_SS_Project_Hours = Sum_Annual_IT_SS_Project_Hours / Sum_TOTAL_Hours * 100
Percentage_Annual_ISO_Project_Hours = Sum_Annual_ISO_Project_Hours / Sum_TOTAL_Hours * 100
Percentage_Annual_Schools_or_Depts_Project_Hours = Sum_Annual_Schools_or_Depts_Project_Hours / Sum_TOTAL_Hours * 100
Percentage_TOTAL_Project_Hours = Sum_TOTAL_Project_Hours / Sum_TOTAL_Hours * 100

csvfile.write("PERCENTAGE,,")
csvfile.write(str(round(Percentage_Total_Time_Away, 2)) + "%,")
csvfile.write(str(round(Percentage_Total_General_Admin, 2)) + "%,")
csvfile.write(str(round(Percentage_Total_Managerial_Admin_Time, 2)) + "%,")
csvfile.write(str(round(Percentage_TOTAL_Admin_Time, 2)) + "%,")
csvfile.write(str(round(Percentage_Total_Support, 2)) + "%,")
csvfile.write(str(round(Percentage_Total_Consulting, 2)) + "%,")
csvfile.write(str(round(Percentage_Total_Other, 2)) + "%,")
csvfile.write(str(round(Percentage_Annual_CI_Project_Hours, 2)) + "%,")
csvfile.write(str(round(Percentage_Annual_AS_Project_Hours, 2)) + "%,")
csvfile.write(str(round(Percentage_Annual_IT_SS_Project_Hours, 2)) + "%,")
csvfile.write(str(round(Percentage_Annual_ISO_Project_Hours, 2)) + "%,")
csvfile.write(str(round(Percentage_Annual_Schools_or_Depts_Project_Hours, 2)) + "%,")
csvfile.write(str(round(Percentage_TOTAL_Project_Hours, 2)))

print()
print(People)
print()
print("If there are errors, most likey there are xlxs files besides the RM Workbooks")
print()
print("Do yourself a favor and only copy the output csv  and chat png files back to Box")
print()

pie_data = [Percentage_TOTAL_Admin_Time, Percentage_Total_Support, Percentage_Total_Consulting, Percentage_TOTAL_Project_Hours]
print (pie_data)
pie_labels = ['Admin_Time', 'Support', 'Consulting', 'Project']
print (pie_labels)

fig1, ax1 = plt.subplots()
ax1.pie(pie_data, labels=pie_labels, autopct='%1.1f%%', startangle=0)
ax1.set_title(team, fontsize=20)
ax1.axis('equal') 
ax1.get_figure().savefig("/DATA/" + team.upper() + "/" + team.lower() + "_pie_chart.png")

